#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Feb 20 17:17:21 2020

@author: aabir
"""
import csv, os, openpyxl,re,string
#This function writes a code that test whether the value in ROW is number 
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
#open the Results excel file
from openpyxl import Workbook
wbr=Workbook()
wsr=wbr.active

rcnt=1; ccnt=1  #row and column count for excel
for csvFilename in os.listdir('.'):     #loop over the csv files
    # skip non-csv files
    if not csvFilename.endswith('.out'):
        continue #continue returns control to the for loop beginning
    regX=re.compile(r'(.*).out')    #to find filename without .out
    xcelN=regX.findall(csvFilename)
    
    csvFileObj=open(csvFilename,newline='')
    readerObj=csv.reader(csvFileObj,delimiter=' ')
    for ROW in readerObj:   #loop over the rows in the current csv file

        if readerObj.line_num==1:   #do nothing for 1st line and remove the header
            print('Removing header from '+csvFilename+'.....')
            continue
        #ignore the transcript file
        if len(ROW)!=2:
            print('Ignoring the FLUENT transcript file: '+csvFilename)
            break        
        
        elif is_number(ROW[0])==False:
            wsr.cell(row=readerObj.line_num,column=ccnt).value=str(xcelN[0]) #add name of the file at the top of excel
            wsr.cell(row=readerObj.line_num+1,column=ccnt).value=ROW[0]
            wsr.cell(row=readerObj.line_num+1,column=ccnt+1).value=ROW[1]
            continue
        elif is_number(ROW[0])==True:
            rcnt+=1
            wsr.cell(row=readerObj.line_num+1,column=ccnt).value=float(ROW[0])
            wsr.cell(row=readerObj.line_num+1,column=ccnt+1).value=float(ROW[1])
    ccnt+=2

wbr.save('results.xlsx')
csvFileObj.close()
wbr.close()
 
    
        
    
        
    
